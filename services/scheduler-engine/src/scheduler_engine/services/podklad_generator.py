"""Generate schedule podkład xlsx from styled templates (28 / 30 / 31 day layouts)."""

from __future__ import annotations

import calendar
import io
from datetime import date
from importlib.resources import files

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

MONTH_UPPER = (
    "STYCZEŃ",
    "LUTY",
    "MARZEC",
    "KWIECIEŃ",
    "MAJ",
    "CZERWIEC",
    "LIPIEC",
    "SIERPIEŃ",
    "WRZESIEŃ",
    "PAŹDZIERNIK",
    "LISTOPAD",
    "GRUDZIEŃ",
)

WEEKDAY_LABELS = ("Niedz.", "Pon.", "Wt.", "Śr.", "Czw.", "Pt.", "Sob.")

# Native templates — preserves merges, styles and formulas from reference xlsx files.
TEMPLATE_BY_DAYS: dict[int, str] = {
    28: "podklad_template_28.xlsx",
    30: "podklad_template_30.xlsx",
    31: "podklad_template_31.xlsx",
}

SHIFT_VALUE_MERGE_ROWS = (10, 12, 14, 16, 18)


def get_days_in_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def format_podklad_file_name(year: int, month: int) -> str:
    days = get_days_in_month(year, month)
    mm = f"{month:02d}"
    last = f"{days:02d}.{mm}"
    return f"PODKŁAD 01.{mm}-{last} R.xlsx"


def weekday_label(year: int, month: int, day: int) -> str:
    weekday = date(year, month, day).weekday()  # Mon=0 … Sun=6
    index = (weekday + 1) % 7  # align with JS Date.getDay() (Sun=0)
    return WEEKDAY_LABELS[index]


def day_weekday_col(day: int) -> int:
    """1-based column for weekday label and day number."""
    return 2 * day + 1


def name_block_col(days_in_month: int) -> int:
    """1-based column for right-side 'nazwisko'."""
    return 2 + 2 * days_in_month + 1


def _template_asset_name(days_in_month: int) -> str:
    if days_in_month in TEMPLATE_BY_DAYS:
        return TEMPLATE_BY_DAYS[days_in_month]
    # 29 days (leap February): start from 30-day layout and drop the last day column pair.
    if days_in_month == 29:
        return TEMPLATE_BY_DAYS[30]
    raise ValueError(f"Unsupported month length: {days_in_month} days")


def _template_path(days_in_month: int):
    return files("scheduler_engine").joinpath("assets", _template_asset_name(days_in_month))


def _unmerge_all(ws: Worksheet) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))


def _apply_merges(ws: Worksheet, days_in_month: int) -> None:
    """Rebuild merges after removing the 30th day column (29-day months only)."""
    name_col = name_block_col(days_in_month)

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=name_col, end_row=2, end_column=name_col + 1)
    ws.cell(1, name_col).value = "=A1"

    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
    ws.merge_cells(start_row=4, start_column=name_col, end_row=5, end_column=name_col)
    ws.merge_cells(start_row=4, start_column=name_col + 1, end_row=5, end_column=name_col + 1)

    for day in range(1, days_in_month + 1):
        col = day_weekday_col(day)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        start = get_column_letter(col)
        end = get_column_letter(col + 1)
        ws.cell(5, col).value = f"={end}4-{start}4"

    for row in SHIFT_VALUE_MERGE_ROWS:
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)


def _adapt_29_day_layout(ws: Worksheet) -> None:
    """Leap-year February: remove day 30 from the 30-day template."""
    _unmerge_all(ws)
    ws.delete_cols(day_weekday_col(30), 2)
    _apply_merges(ws, 29)


def _fill_calendar(ws: Worksheet, year: int, month: int, days_in_month: int) -> None:
    title = f"{MONTH_UPPER[month - 1]} {year}"
    ws.cell(1, 1).value = title

    for day in range(1, days_in_month + 1):
        col = day_weekday_col(day)
        ws.cell(1, col).value = weekday_label(year, month, day)
        ws.cell(2, col).value = day

    ws.cell(3, 1).value = "nazwisko"
    ws.cell(3, 2).value = "imię"

    name_col = name_block_col(days_in_month)
    ws.cell(3, name_col).value = "nazwisko"
    ws.cell(3, name_col + 1).value = "imię"


def generate_podklad_workbook(year: int, month: int):
    if month < 1 or month > 12:
        raise ValueError("month must be 1–12")
    if year < 2000 or year > 2100:
        raise ValueError("year out of range")

    days_in_month = get_days_in_month(year, month)

    with _template_path(days_in_month).open("rb") as template_file:
        workbook = load_workbook(template_file)
    worksheet = workbook.active

    if days_in_month == 29:
        _adapt_29_day_layout(worksheet)

    _fill_calendar(worksheet, year, month, days_in_month)
    return workbook


def generate_podklad_bytes(year: int, month: int) -> bytes:
    workbook = generate_podklad_workbook(year, month)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
