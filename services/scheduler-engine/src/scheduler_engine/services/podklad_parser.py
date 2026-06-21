"""Parse worker podkład xlsx files into structured availability data."""

from __future__ import annotations

import base64
import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from typing import Literal

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from scheduler_engine.services.podklad_colors import FillKind, classify_pair_fill
from scheduler_engine.services.podklad_generator import day_weekday_col, get_days_in_month
from scheduler_engine.services.podklad_layout import (
    AFTERNOON_COLOR_ROW,
    AFTERNOON_INTERVAL_ROW,
    CUSTOM_INTERVAL_COLOR_ROW,
    CUSTOM_INTERVAL_VALUE_ROW,
    DEFAULT_AFTERNOON_RANGE,
    DEFAULT_FULL_DAY_RANGE,
    DEFAULT_MORNING_RANGE,
    FLEXIBLE_COLOR_ROW,
    FLEXIBLE_INTERVAL_ROW,
    MORNING_COLOR_ROW,
    MORNING_INTERVAL_ROW,
    NAME_HEADER_ROW_SCAN_LIMIT,
)
from scheduler_engine.services.time_range import (
    TimeRange,
    merge_time_ranges,
    parse_hour_pair,
    parse_time_range_text,
)

WorkerRole = Literal["worker", "boss"]

POLISH_MONTHS: dict[str, int] = {
    "STYCZEN": 1,
    "LUTY": 2,
    "MARZEC": 3,
    "KWIECIEN": 4,
    "MAJ": 5,
    "CZERWIEC": 6,
    "LIPIEC": 7,
    "SIERPIEN": 8,
    "WRZESIEN": 9,
    "PAZDZIERNIK": 10,
    "LISTOPAD": 11,
    "GRUDZIEN": 12,
}


@dataclass(slots=True)
class DayDisposition:
    date: str
    ranges: list[TimeRange] = field(default_factory=list)
    morning_color: FillKind = "none"
    afternoon_color: FillKind = "none"

    @property
    def available(self) -> bool:
        return len(self.ranges) > 0

    def to_json(self) -> dict[str, object]:
        return {
            "date": self.date,
            "ranges": [{"start": start, "end": end} for start, end in self.ranges],
            "morningColor": self.morning_color,
            "afternoonColor": self.afternoon_color,
        }


@dataclass(slots=True)
class ParsedWorkerDraft:
    worker_id: str
    draft_id: str
    file_name: str
    first_name: str
    last_name: str
    role: WorkerRole
    year: int
    month: int
    available_as_worker: bool = True
    days: list[DayDisposition] = field(default_factory=list)

    def to_json(self) -> dict[str, object]:
        return {
            "workerId": self.worker_id,
            "draftId": self.draft_id,
            "fileName": self.file_name,
            "firstName": self.first_name,
            "lastName": self.last_name,
            "role": self.role,
            "year": self.year,
            "month": self.month,
            "availableAsWorker": self.available_as_worker,
            "days": [day.to_json() for day in self.days],
        }


def parse_worker_draft(
    *,
    worker_id: str,
    draft_id: str,
    file_name: str,
    content_base64: str,
    role: WorkerRole,
    year: int,
    month: int,
) -> ParsedWorkerDraft:
    file_bytes = base64.b64decode(content_base64)
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    if worksheet is None:
        raise ValueError(f"Plik {file_name} nie zawiera arkusza")

    parsed_year, parsed_month = _extract_year_month(worksheet, file_name, year, month)
    first_name, last_name = _extract_worker_names(worksheet)
    days = _extract_day_dispositions(worksheet, parsed_year, parsed_month)

    return ParsedWorkerDraft(
        worker_id=worker_id,
        draft_id=draft_id,
        file_name=file_name,
        first_name=first_name,
        last_name=last_name,
        role=role,
        year=parsed_year,
        month=parsed_month,
        days=days,
    )


def _extract_year_month(
    worksheet: Worksheet,
    file_name: str,
    fallback_year: int,
    fallback_month: int,
) -> tuple[int, int]:
    title = _cell_text(worksheet, 0, 0)
    from_title = _parse_title_year_month(title)
    if from_title:
        return from_title

    from_name = _parse_file_name_year_month(file_name)
    if from_name:
        return from_name

    return fallback_year, fallback_month


def _extract_worker_names(worksheet: Worksheet) -> tuple[str, str]:
    max_row = min(worksheet.max_row or NAME_HEADER_ROW_SCAN_LIMIT, NAME_HEADER_ROW_SCAN_LIMIT)

    for row in range(max_row):
        last_name_header = _cell_text(worksheet, row, 0)
        first_name_header = _cell_text(worksheet, row, 1)
        if not _is_last_name_header(last_name_header) or not _is_first_name_header(first_name_header):
            continue

        last_name = _cell_text(worksheet, row + 1, 0)
        first_name = _cell_text(worksheet, row + 1, 1)
        if len(last_name) >= 2 and len(first_name) >= 2:
            return first_name, last_name

        raise ValueError("Brak imienia i nazwiska w podkładzie")

    raise ValueError("Nie znaleziono nagłówków nazwisko/imię w podkładzie")


def _extract_day_dispositions(
    worksheet: Worksheet,
    year: int,
    month: int,
) -> list[DayDisposition]:
    days_in_month = get_days_in_month(year, month)
    result: list[DayDisposition] = []

    for day in range(1, days_in_month + 1):
        date = f"{year}-{month:02d}-{day:02d}"
        col = day_weekday_col(day)
        ranges: list[TimeRange] = []

        morning, morning_color = _resolve_band_range(
            worksheet,
            col,
            color_row=MORNING_COLOR_ROW,
            interval_row=MORNING_INTERVAL_ROW,
            color_default=DEFAULT_MORNING_RANGE,
            accepted_colors={"yellow", "white"},
            preview_color="yellow",
        )
        if morning:
            ranges.append(morning)

        afternoon, afternoon_color = _resolve_band_range(
            worksheet,
            col,
            color_row=AFTERNOON_COLOR_ROW,
            interval_row=AFTERNOON_INTERVAL_ROW,
            color_default=DEFAULT_AFTERNOON_RANGE,
            accepted_colors={"purple", "white"},
            preview_color="purple",
        )
        if afternoon:
            ranges.append(afternoon)

        flexible, _flex_color = _resolve_band_range(
            worksheet,
            col,
            color_row=FLEXIBLE_COLOR_ROW,
            interval_row=FLEXIBLE_INTERVAL_ROW,
            color_default=DEFAULT_FULL_DAY_RANGE,
            accepted_colors={"white"},
            preview_color="yellow",
        )
        if flexible:
            ranges.append(flexible)
            if morning_color == "none":
                morning_color = "yellow"
            if afternoon_color == "none":
                afternoon_color = "purple"

        custom, custom_color = _resolve_band_range(
            worksheet,
            col,
            color_row=CUSTOM_INTERVAL_COLOR_ROW,
            interval_row=CUSTOM_INTERVAL_VALUE_ROW,
            color_default=None,
            accepted_colors={"yellow", "purple", "white"},
            require_custom_or_color=True,
            preview_color="yellow",
        )
        if custom:
            ranges.append(custom)
            if custom_color == "yellow" and morning_color == "none":
                morning_color = "yellow"
            if custom_color == "purple" and afternoon_color == "none":
                afternoon_color = "purple"
            if custom_color == "white":
                if morning_color == "none":
                    morning_color = "yellow"
                if afternoon_color == "none":
                    afternoon_color = "purple"

        result.append(
            DayDisposition(
                date=date,
                ranges=merge_time_ranges(ranges),
                morning_color=morning_color,
                afternoon_color=afternoon_color,
            )
        )

    return result


def _resolve_band_range(
    worksheet: Worksheet,
    col: int,
    *,
    color_row: int,
    interval_row: int,
    color_default: TimeRange | None,
    accepted_colors: set[FillKind],
    require_custom_or_color: bool = False,
    preview_color: FillKind = "none",
) -> tuple[TimeRange | None, FillKind]:
    left = worksheet.cell(color_row, col)
    right = worksheet.cell(color_row, col + 1)
    fill_kind = classify_pair_fill(left, right)

    custom = _parse_interval_from_pair(worksheet, color_row, interval_row, col)
    if custom:
        display_color = fill_kind if fill_kind in {"yellow", "purple", "white"} else preview_color
        if display_color == "white":
            display_color = preview_color
        return custom, display_color

    if fill_kind == "none":
        return None, "none"

    if fill_kind not in accepted_colors:
        return None, "none"

    if fill_kind == "white":
        return DEFAULT_FULL_DAY_RANGE, preview_color

    if color_default is None:
        return None if require_custom_or_color else None, "none"

    display_color = fill_kind if fill_kind in {"yellow", "purple"} else preview_color
    return color_default, display_color


def _parse_interval_from_pair(
    worksheet: Worksheet,
    color_row: int,
    interval_row: int,
    col: int,
) -> TimeRange | None:
    for row in (color_row, interval_row):
        left_value = worksheet.cell(row, col).value
        right_value = worksheet.cell(row, col + 1).value

        hour_pair = parse_hour_pair(left_value, right_value)
        if hour_pair:
            return hour_pair

        for value in (left_value, right_value):
            if isinstance(value, str):
                parsed = parse_time_range_text(value)
                if parsed:
                    return parsed

    return None


def _cell_text(worksheet: Worksheet, row: int, col: int) -> str:
    value = worksheet.cell(row + 1, col + 1).value
    if value is None:
        return ""
    return str(value).strip()


def _normalize_token(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value.strip().upper())
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def _is_last_name_header(value: str) -> bool:
    return _normalize_token(value) == "NAZWISKO"


def _is_first_name_header(value: str) -> bool:
    return _normalize_token(value).startswith("IMIE")


def _parse_title_year_month(title: str) -> tuple[int, int] | None:
    match = re.match(r"^([A-Za-zĄĆĘŁŃÓŚŹŻąćęłńóśźż]+)\s+(20\d{2})$", title.strip())
    if not match:
        return None

    month = POLISH_MONTHS.get(_normalize_token(match.group(1)))
    year = int(match.group(2))
    if month is None or year < 2000 or year > 2100:
        return None
    return year, month


def _parse_file_name_year_month(file_name: str) -> tuple[int, int] | None:
    base = re.sub(r"\.(xlsx|xls)$", "", file_name, flags=re.IGNORECASE)
    patterns = [
        r"(?:^|[^\d])(20\d{2})[-_.]([01]?\d{1,2})(?:[^\d]|$)",
        r"(?:^|[^\d])([01]?\d{1,2})[-_.](20\d{2})(?:[^\d]|$)",
        r"(?:^|[^\d])01\.([01]?\d{1,2})-",
    ]

    for pattern in patterns:
        match = re.search(pattern, base)
        if not match:
            continue

        if "01." in pattern:
            month = int(match.group(1))
            year_match = re.search(r"(20\d{2})", base)
            year = int(year_match.group(1)) if year_match else 0
        else:
            year = int(match.group(1) if len(match.group(1)) == 4 else match.group(2))
            month = int(match.group(2) if len(match.group(1)) == 4 else match.group(1))

        if 1 <= month <= 12 and 2000 <= year <= 2100:
            return year, month

    return None
