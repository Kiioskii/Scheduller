from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from scheduler_engine.services.podklad_generator import (
    day_weekday_col,
    format_podklad_file_name,
    generate_podklad_bytes,
    get_days_in_month,
    name_block_col,
    weekday_label,
)


def test_format_podklad_file_name() -> None:
    assert format_podklad_file_name(2026, 6) == "PODKŁAD 01.06-30.06 R.xlsx"
    assert get_days_in_month(2026, 2) == 28
    assert format_podklad_file_name(2026, 2) == "PODKŁAD 01.02-28.02 R.xlsx"


def test_weekday_label_matches_js_convention() -> None:
    assert weekday_label(2026, 6, 1) == "Pon."


def test_june_layout_matches_template_structure() -> None:
    content = generate_podklad_bytes(2026, 6)
    ws = load_workbook(BytesIO(content)).active

    assert ws.cell(1, 1).value == "CZERWIEC 2026"
    assert ws.cell(1, 3).value == "Pon."
    assert ws.cell(2, 3).value == 1
    assert ws.cell(2, 61).value == 30
    assert ws.cell(3, 1).value == "nazwisko"
    assert ws.cell(3, 2).value == "imię"
    assert ws.cell(9, 2).value == "RANO"
    assert ws.cell(11, 2).value == "POPOŁUDNIE"
    assert ws.cell(15, 2).value == "PRZEDZIAŁ"
    assert ws.cell(15, 4).value == 13
    assert ws.cell(15, 5).value == 22
    assert len(ws.merged_cells.ranges) >= 35


def test_february_has_28_days() -> None:
    content = generate_podklad_bytes(2026, 2)
    ws = load_workbook(BytesIO(content)).active
    days = get_days_in_month(2026, 2)
    name_col = name_block_col(days)

    assert ws.cell(1, 1).value == "LUTY 2026"
    assert ws.cell(2, day_weekday_col(28)).value == 28
    assert ws.cell(2, day_weekday_col(29)).value is None
    assert ws.cell(3, name_col).value == "nazwisko"
    assert ws.cell(3, name_col + 1).value == "imię"
    assert ws.cell(1, name_col).value == "=A1"
    col28 = day_weekday_col(28)
    expected_formula = f"={get_column_letter(col28 + 1)}4-{get_column_letter(col28)}4"
    assert ws.cell(5, col28).value == expected_formula


def test_leap_february_has_29_days() -> None:
    content = generate_podklad_bytes(2028, 2)
    ws = load_workbook(BytesIO(content)).active

    assert get_days_in_month(2028, 2) == 29
    assert ws.cell(2, day_weekday_col(29)).value == 29
    assert ws.cell(2, day_weekday_col(30)).value is None


def test_january_has_31_days() -> None:
    content = generate_podklad_bytes(2026, 1)
    ws = load_workbook(BytesIO(content)).active
    days = get_days_in_month(2026, 1)
    name_col = name_block_col(days)

    assert ws.cell(1, 1).value == "STYCZEŃ 2026"
    assert ws.cell(2, day_weekday_col(31)).value == 31
    assert ws.cell(1, day_weekday_col(31)).value == "Sob."
    assert ws.cell(3, name_col).value == "nazwisko"
    assert ws.cell(1, name_col).value == "=A1"
    assert len(ws.merged_cells.ranges) >= 38


def _has_black_fill(cell) -> bool:
    fill = cell.fill
    rgb = getattr(fill.fgColor, "rgb", "") or ""
    rgb_value = str(rgb)
    return fill.patternType == "solid" and rgb_value.endswith("000000")


def test_holiday_days_are_marked_in_row_4() -> None:
    content = generate_podklad_bytes(2026, 6, ["2026-06-01", "2026-06-15"])
    ws = load_workbook(BytesIO(content)).active

    holiday_col = day_weekday_col(1)
    regular_col = day_weekday_col(2)

    for col in (holiday_col, holiday_col + 1):
        assert _has_black_fill(ws.cell(4, col))

    assert not _has_black_fill(ws.cell(4, regular_col))
    assert not _has_black_fill(ws.cell(5, holiday_col))
    assert not _has_black_fill(ws.cell(10, holiday_col))
