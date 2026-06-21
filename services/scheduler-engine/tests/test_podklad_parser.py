import base64
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from scheduler_engine.services.podklad_generator import day_weekday_col, generate_podklad_bytes
from scheduler_engine.services.podklad_parser import parse_worker_draft

YELLOW_FILL = PatternFill(patternType="solid", fgColor="FFFFFF99")
PURPLE_FILL = PatternFill(patternType="solid", fgColor="FFCCCCFF")
WHITE_FILL = PatternFill(patternType="solid", fgColor="FFFFFFFF")


def _filled_podklad_base64() -> str:
    content = generate_podklad_bytes(2026, 6)
    workbook = load_workbook(BytesIO(content))
    worksheet = workbook.active
    worksheet.cell(4, 1).value = "Kijowski"
    worksheet.cell(4, 2).value = "Maciej"

    day_one_col = day_weekday_col(5)
    worksheet.cell(9, day_one_col + 1).fill = YELLOW_FILL

    day_two_col = day_weekday_col(6)
    worksheet.cell(11, day_two_col).fill = PURPLE_FILL

    day_three_col = day_weekday_col(7)
    worksheet.cell(13, day_three_col + 1).fill = WHITE_FILL

    day_four_col = day_weekday_col(8)
    worksheet.cell(9, day_four_col + 1).fill = YELLOW_FILL
    worksheet.cell(9, day_four_col).value = "10:00-14:00"

    buffer = BytesIO()
    workbook.save(buffer)
    return base64.b64encode(buffer.getvalue()).decode("ascii")


def test_parse_worker_draft_extracts_color_based_ranges() -> None:
    parsed = parse_worker_draft(
        worker_id="42",
        draft_id="10",
        file_name="PODKŁAD 01.06-30.06 R.xlsx",
        content_base64=_filled_podklad_base64(),
        role="worker",
        year=2026,
        month=6,
    )

    assert parsed.first_name == "Maciej"
    assert parsed.last_name == "Kijowski"

    day_one = next(day for day in parsed.days if day.date == "2026-06-05")
    day_two = next(day for day in parsed.days if day.date == "2026-06-06")
    day_three = next(day for day in parsed.days if day.date == "2026-06-07")
    day_four = next(day for day in parsed.days if day.date == "2026-06-08")

    assert day_one.ranges == [("08:00", "15:15")]
    assert day_two.ranges == [("15:00", "22:00")]
    assert day_three.ranges == [("08:00", "22:00")]
    assert day_four.ranges == [("10:00", "14:00")]

    payload = parsed.to_json()
    assert payload["days"][0]["ranges"][0]["start"] == "08:00"
